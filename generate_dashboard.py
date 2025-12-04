#!/usr/bin/env python3
"""
SMU ELDIS Radar Dashboard Generator
====================================

 ███████╗██╗     ██████╗ ██╗███████╗    ██████╗  █████╗ ██████╗  █████╗ ██████╗ 
 ██╔════╝██║     ██╔══██╗██║██╔════╝    ██╔══██╗██╔══██╗██╔══██╗██╔══██╗██╔══██╗
 █████╗  ██║     ██║  ██║██║███████╗    ██████╔╝███████║██║  ██║███████║██████╔╝
 ██╔══╝  ██║     ██║  ██║██║╚════██║    ██╔══██╗██╔══██║██║  ██║██╔══██║██╔══██╗
 ███████╗███████╗██████╔╝██║███████║    ██║  ██║██║  ██║██████╔╝██║  ██║██║  ██║
 ╚══════╝╚══════╝╚═════╝ ╚═╝╚══════╝    ╚═╝  ╚═╝╚═╝  ╚═╝╚═════╝ ╚═╝  ╚═╝╚═╝  ╚═╝

A professional Excel dashboard generator for SMU ELDIS Radar operations management.
Generates an Excel workbook with routine task tracking and GoI-style project lifecycles.

Author: SMU Dashboard Team
License: MIT
"""

from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# =============================================================================
# Configuration Constants
# =============================================================================

OUTPUT_FILE = "SMU_Lite_Dashboard.xlsx"

# Color scheme for styling
COLORS = {
    "header_bg": "1F4E79",       # Dark blue
    "header_text": "FFFFFF",    # White
    "overdue_bg": "FF6B6B",     # Red for overdue
    "warning_bg": "FFD93D",     # Yellow for due soon
    "completed_bg": "6BCB77",   # Green for completed
    "alt_row": "E8F4F8",        # Light blue for alternating rows
}

# Styling objects
HEADER_FONT = Font(bold=True, color=COLORS["header_text"], size=11)
HEADER_FILL = PatternFill(start_color=COLORS["header_bg"], 
                          end_color=COLORS["header_bg"], 
                          fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Conditional formatting fills
RED_FILL = PatternFill(start_color=COLORS["overdue_bg"], 
                       end_color=COLORS["overdue_bg"], 
                       fill_type="solid")
YELLOW_FILL = PatternFill(start_color=COLORS["warning_bg"], 
                          end_color=COLORS["warning_bg"], 
                          fill_type="solid")
GREEN_FILL = PatternFill(start_color=COLORS["completed_bg"], 
                         end_color=COLORS["completed_bg"], 
                         fill_type="solid")


# =============================================================================
# Data: Routine Actions (Sl.No 43-51)
# =============================================================================

ROUTINE_DATA = [
    {
        "Sl.No": 43,
        "Action Point": "Finalization of policy for routine optimization of RADARs",
        "Responsible": "GM(CNS-Sur)/ED(CNS-O&M)",
        "Accountable": "ED(CNS-O&M)",
        "Open Date": date(2025, 1, 15),
        "Due Date": date(2025, 5, 22),
        "Status": "Open",
        "Priority": "Medium",
        "SLA": "Routine",
        "Actions Taken": "Policy already approved",
        "Remarks": ""
    },
    {
        "Sl.No": 44,
        "Action Point": "Expansion of in-house repair capabilities for ADS-B sensors, ERA & Sensis ASMGCS sensors",
        "Responsible": "GM(CMC)/GM(CNS-P1)",
        "Accountable": "GM(CMC)",
        "Open Date": date(2025, 1, 20),
        "Due Date": date(2025, 5, 22),
        "Status": "Open",
        "Priority": "High",
        "SLA": "Strategic",
        "Actions Taken": "1. ADS-B coordination to be initiated; 2. Test jig not available",
        "Remarks": "SURV SMU declare manpower/skills/tools; need ASMGCS test jig"
    },
    {
        "Sl.No": 45,
        "Action Point": "Advance Factory Training Policy for SMU executives",
        "Responsible": "In-Charge ELDIS SMU/GM(CMC)",
        "Accountable": "GM(CMC)",
        "Open Date": date(2025, 2, 1),
        "Due Date": date(2025, 5, 22),
        "Status": "Open",
        "Priority": "High",
        "SLA": "Strategic",
        "Actions Taken": "Matter taken up with CNSOM, no policy finalized",
        "Remarks": "New file for specialist FT requirement"
    },
    {
        "Sl.No": 46,
        "Action Point": "Policy for Dedicated Vehicles for SMUs",
        "Responsible": "GM(CMC)",
        "Accountable": "GM(CMC)",
        "Open Date": date(2025, 2, 10),
        "Due Date": date(2025, 5, 22),
        "Status": "Open",
        "Priority": "Medium",
        "SLA": "Routine",
        "Actions Taken": "Proposal to be initiated",
        "Remarks": "Efile to CMC"
    },
    {
        "Sl.No": 47,
        "Action Point": "Establishment of Test System Facility at SMU A-SMGCS",
        "Responsible": "In-Charge ASMGCS SMU/GM(CMC)",
        "Accountable": "GM(CMC)",
        "Open Date": date(2025, 2, 15),
        "Due Date": date(2025, 5, 22),
        "Status": "In Progress",
        "Priority": "High",
        "SLA": "Project",
        "Actions Taken": "Procurement under process",
        "Remarks": ""
    },
    {
        "Sl.No": 48,
        "Action Point": "In-House ITC (Installation, Testing & Commissioning) for RADARs",
        "Responsible": "GM(CMC)/All ED(CNS)",
        "Accountable": "ED(CNS)",
        "Open Date": date(2025, 3, 1),
        "Due Date": date(2025, 5, 22),
        "Status": "Open",
        "Priority": "High",
        "SLA": "Project",
        "Actions Taken": "Coordination done; proposal pending",
        "Remarks": ""
    },
    {
        "Sl.No": 49,
        "Action Point": "Provision of broadband & firewall at A-SMGCS stations",
        "Responsible": "GM(CMC)",
        "Accountable": "GM(CMC)",
        "Open Date": date(2025, 3, 10),
        "Due Date": date(2025, 5, 22),
        "Status": "Open",
        "Priority": "Medium",
        "SLA": "Routine",
        "Actions Taken": "Stations advised to install broadband",
        "Remarks": "Firewall by ASMGCS after tech discussion"
    },
    {
        "Sl.No": 50,
        "Action Point": "Standard packing & insurance for spares sent to SMU",
        "Responsible": "GM(CMC)",
        "Accountable": "GM(CMC)",
        "Open Date": date(2025, 3, 15),
        "Due Date": date(2025, 5, 22),
        "Status": "In Progress",
        "Priority": "Medium",
        "SLA": "Routine",
        "Actions Taken": "Guidelines to be formulated",
        "Remarks": "CMC circular based on SMU proposal"
    },
    {
        "Sl.No": 51,
        "Action Point": "Surveillance SMU rep in CNS/ATM Automation spec committees",
        "Responsible": "GM(CMC)/GM(CNS-Aut)/ED(O&M)",
        "Accountable": "ED(O&M)",
        "Open Date": date(2025, 3, 20),
        "Due Date": date(2025, 5, 22),
        "Status": "Open",
        "Priority": "Medium",
        "SLA": "Strategic",
        "Actions Taken": "To be initiated",
        "Remarks": "ITWP + Automation tracking issues"
    }
]


# =============================================================================
# Data: Projects (GoI-style Tender Lifecycle)
# =============================================================================

PROJECT_DATA = [
    {
        "Project ID": "P-01",
        "Project Title": "ASMGCS Test System Procurement",
        "Stage": "Procurement",
        "Start Date": date(2025, 1, 15),
        "Target Date": date(2025, 6, 30),
        "Actual Completion": None,
        "Status": "In Progress",
        "Owner": "GM(CMC)",
        "Remarks": "Procurement process initiated; vendor evaluation ongoing"
    },
    {
        "Project ID": "P-02",
        "Project Title": "ADS-B In-house Repair Capability",
        "Stage": "Tech Validation",
        "Start Date": date(2025, 2, 1),
        "Target Date": date(2025, 8, 15),
        "Actual Completion": None,
        "Status": "In Progress",
        "Owner": "GM(CNS-P1)",
        "Remarks": "Technical validation in progress; awaiting test equipment"
    },
    {
        "Project ID": "P-03",
        "Project Title": "Radar ITC In-house Implementation",
        "Stage": "Proposal",
        "Start Date": date(2025, 3, 1),
        "Target Date": date(2025, 9, 30),
        "Actual Completion": None,
        "Status": "Open",
        "Owner": "ED(CNS)",
        "Remarks": "Proposal under preparation; coordination with multiple EDs"
    }
]


# =============================================================================
# Data: Lists for Data Validation
# =============================================================================

LISTS_DATA = {
    "Status": ["Open", "In Progress", "Completed", "Blocked"],
    "Priority": ["High", "Medium", "Low"],
    "SLA": ["Routine", "Strategic", "Project", "Emergency"],
    "Project Stages": [
        "Concept", "Feasibility", "AA (Admin Approval)", 
        "FC (Financial Concurrence)", "RFP", "Bidding", 
        "Evaluation", "Award", "Execution", "Testing", "Closure"
    ]
}


# =============================================================================
# Helper Functions
# =============================================================================

def apply_header_style(ws, row_num, start_col=1, end_col=None):
    """Apply professional header styling to a row."""
    if end_col is None:
        end_col = ws.max_column
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def apply_borders(ws, start_row, end_row, start_col, end_col):
    """Apply thin borders to a range of cells."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def auto_fit_columns(ws, min_width=8, max_width=50):
    """Auto-fit column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            except (TypeError, AttributeError):
                pass
        
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column].width = adjusted_width


def get_today():
    """Get today's date."""
    return date.today()


# =============================================================================
# Sheet Creation Functions
# =============================================================================

def create_lists_sheet(wb):
    """
    Create the Lists sheet for data validation dropdowns.
    This sheet stores reference lists for Status, Priority, SLA, and Project Stages.
    """
    ws = wb.create_sheet("Lists")
    
    # Write headers and data for each list
    col = 1
    for list_name, values in LISTS_DATA.items():
        # Header
        ws.cell(row=1, column=col, value=list_name)
        apply_header_style(ws, 1, col, col)
        
        # Values
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col, value=value)
            ws.cell(row=row_idx, column=col).border = THIN_BORDER
        
        col += 1
    
    auto_fit_columns(ws)
    return ws


def create_routine_table(wb):
    """
    Create the Routine_Table sheet with SMU action items.
    Includes calculated columns for Days Remaining, Days Open, and Overdue Flag.
    """
    ws = wb.create_sheet("Routine_Table")
    today = get_today()
    
    # Define columns (including calculated columns)
    columns = [
        "Sl.No", "Action Point", "Responsible", "Accountable", 
        "Open Date", "Due Date", "Status", "Priority", "SLA", 
        "Actions Taken", "Remarks", 
        "Days Remaining", "Days Open", "Overdue Flag"
    ]
    
    # Write headers
    for col_idx, header in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    apply_header_style(ws, 1, 1, len(columns))
    
    # Write data rows
    for row_idx, item in enumerate(ROUTINE_DATA, start=2):
        ws.cell(row=row_idx, column=1, value=item["Sl.No"])
        ws.cell(row=row_idx, column=2, value=item["Action Point"])
        ws.cell(row=row_idx, column=3, value=item["Responsible"])
        ws.cell(row=row_idx, column=4, value=item["Accountable"])
        ws.cell(row=row_idx, column=5, value=item["Open Date"])
        ws.cell(row=row_idx, column=6, value=item["Due Date"])
        ws.cell(row=row_idx, column=7, value=item["Status"])
        ws.cell(row=row_idx, column=8, value=item["Priority"])
        ws.cell(row=row_idx, column=9, value=item["SLA"])
        ws.cell(row=row_idx, column=10, value=item["Actions Taken"])
        ws.cell(row=row_idx, column=11, value=item["Remarks"])
        
        # Calculated columns
        open_date = item["Open Date"]
        due_date = item["Due Date"]
        days_remaining = (due_date - today).days
        # Days Open: Today - Open Date (how long the item has been open)
        days_open = (today - open_date).days if item["Status"] != "Completed" else 0
        
        ws.cell(row=row_idx, column=12, value=days_remaining)
        ws.cell(row=row_idx, column=13, value=days_open)
        
        # Overdue Flag
        if days_remaining < 0 and item["Status"] != "Completed":
            ws.cell(row=row_idx, column=14, value="OVERDUE")
        else:
            ws.cell(row=row_idx, column=14, value="")
        
        # Format date columns
        ws.cell(row=row_idx, column=5).number_format = "YYYY-MM-DD"
        ws.cell(row=row_idx, column=6).number_format = "YYYY-MM-DD"
    
    # Apply borders to data area
    apply_borders(ws, 1, len(ROUTINE_DATA) + 1, 1, len(columns))
    
    # Add Data Validation for Status, Priority, SLA columns
    status_dv = DataValidation(
        type="list",
        formula1='"Open,In Progress,Completed,Blocked"',
        allow_blank=True
    )
    status_dv.error = "Please select a valid Status"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)
    status_dv.add(f"G2:G{len(ROUTINE_DATA) + 1}")
    
    priority_dv = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True
    )
    priority_dv.error = "Please select a valid Priority"
    priority_dv.errorTitle = "Invalid Priority"
    ws.add_data_validation(priority_dv)
    priority_dv.add(f"H2:H{len(ROUTINE_DATA) + 1}")
    
    sla_dv = DataValidation(
        type="list",
        formula1='"Routine,Strategic,Project,Emergency"',
        allow_blank=True
    )
    sla_dv.error = "Please select a valid SLA"
    sla_dv.errorTitle = "Invalid SLA"
    ws.add_data_validation(sla_dv)
    sla_dv.add(f"I2:I{len(ROUTINE_DATA) + 1}")
    
    # Conditional Formatting
    last_row = len(ROUTINE_DATA) + 1
    
    # Red for overdue items (Days Remaining < 0 AND Status != Completed)
    red_rule = FormulaRule(
        formula=['AND($L2<0,$G2<>"Completed")'],
        fill=RED_FILL
    )
    ws.conditional_formatting.add(f"A2:N{last_row}", red_rule)
    
    # Yellow for items due within 7 days (Days Remaining 0-7 AND Status != Completed)
    yellow_rule = FormulaRule(
        formula=['AND($L2>=0,$L2<=7,$G2<>"Completed")'],
        fill=YELLOW_FILL
    )
    ws.conditional_formatting.add(f"A2:N{last_row}", yellow_rule)
    
    # Green for completed items
    green_rule = FormulaRule(
        formula=['$G2="Completed"'],
        fill=GREEN_FILL
    )
    ws.conditional_formatting.add(f"A2:N{last_row}", green_rule)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    auto_fit_columns(ws)
    return ws


def create_project_table(wb):
    """
    Create the Project_Table sheet for GoI-style tender lifecycle tracking.
    Includes calculated columns for Delay Days and Days to Target.
    """
    ws = wb.create_sheet("Project_Table")
    today = get_today()
    
    # Define columns
    columns = [
        "Project ID", "Project Title", "Stage", "Start Date", 
        "Target Date", "Actual Completion", "Status", "Owner",
        "Delay Days", "Days to Target", "Remarks"
    ]
    
    # Write headers
    for col_idx, header in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    apply_header_style(ws, 1, 1, len(columns))
    
    # Write data rows
    for row_idx, project in enumerate(PROJECT_DATA, start=2):
        ws.cell(row=row_idx, column=1, value=project["Project ID"])
        ws.cell(row=row_idx, column=2, value=project["Project Title"])
        ws.cell(row=row_idx, column=3, value=project["Stage"])
        ws.cell(row=row_idx, column=4, value=project["Start Date"])
        ws.cell(row=row_idx, column=5, value=project["Target Date"])
        ws.cell(row=row_idx, column=6, value=project["Actual Completion"])
        ws.cell(row=row_idx, column=7, value=project["Status"])
        ws.cell(row=row_idx, column=8, value=project["Owner"])
        
        # Calculated columns
        target_date = project["Target Date"]
        actual_completion = project["Actual Completion"]
        
        # Delay Days: If completed, compare actual vs target; otherwise days past target if overdue
        if actual_completion:
            delay_days = (actual_completion - target_date).days
        elif today > target_date:
            delay_days = (today - target_date).days
        else:
            delay_days = 0
        
        # Days to Target: Target Date - Today (negative if past)
        days_to_target = (target_date - today).days
        
        ws.cell(row=row_idx, column=9, value=delay_days)
        ws.cell(row=row_idx, column=10, value=days_to_target)
        ws.cell(row=row_idx, column=11, value=project["Remarks"])
        
        # Format date columns
        ws.cell(row=row_idx, column=4).number_format = "YYYY-MM-DD"
        ws.cell(row=row_idx, column=5).number_format = "YYYY-MM-DD"
        if actual_completion:
            ws.cell(row=row_idx, column=6).number_format = "YYYY-MM-DD"
    
    # Apply borders to data area
    apply_borders(ws, 1, len(PROJECT_DATA) + 1, 1, len(columns))
    
    # Add Data Validation for Status column
    status_dv = DataValidation(
        type="list",
        formula1='"Open,In Progress,Completed,Blocked"',
        allow_blank=True
    )
    status_dv.error = "Please select a valid Status"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)
    status_dv.add(f"G2:G{len(PROJECT_DATA) + 1}")
    
    # Add Data Validation for Stage column
    stages = ",".join(LISTS_DATA["Project Stages"])
    stage_dv = DataValidation(
        type="list",
        formula1=f'"{stages}"',
        allow_blank=True
    )
    stage_dv.error = "Please select a valid Stage"
    stage_dv.errorTitle = "Invalid Stage"
    ws.add_data_validation(stage_dv)
    stage_dv.add(f"C2:C{len(PROJECT_DATA) + 1}")
    
    # Conditional Formatting
    last_row = len(PROJECT_DATA) + 1
    
    # Red for overdue projects (Days to Target < 0 AND Status != Completed)
    red_rule = FormulaRule(
        formula=['AND($J2<0,$G2<>"Completed")'],
        fill=RED_FILL
    )
    ws.conditional_formatting.add(f"A2:K{last_row}", red_rule)
    
    # Yellow for projects due within 30 days
    yellow_rule = FormulaRule(
        formula=['AND($J2>=0,$J2<=30,$G2<>"Completed")'],
        fill=YELLOW_FILL
    )
    ws.conditional_formatting.add(f"A2:K{last_row}", yellow_rule)
    
    # Green for completed projects
    green_rule = FormulaRule(
        formula=['$G2="Completed"'],
        fill=GREEN_FILL
    )
    ws.conditional_formatting.add(f"A2:K{last_row}", green_rule)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    auto_fit_columns(ws)
    return ws


def create_control_panel(wb):
    """
    Create the Control_Panel (Dashboard) sheet with KPI summary metrics.
    This is a one-page summary of all actions and their status.
    """
    ws = wb.create_sheet("Control_Panel", 0)  # Insert at the beginning
    today = get_today()
    
    # Calculate metrics from ROUTINE_DATA
    total_actions = len(ROUTINE_DATA)
    open_count = sum(1 for item in ROUTINE_DATA if item["Status"] == "Open")
    in_progress_count = sum(1 for item in ROUTINE_DATA if item["Status"] == "In Progress")
    completed_count = sum(1 for item in ROUTINE_DATA if item["Status"] == "Completed")
    
    # Overdue: Days Remaining < 0 AND Status != Completed
    overdue_count = sum(
        1 for item in ROUTINE_DATA 
        if (item["Due Date"] - today).days < 0 and item["Status"] != "Completed"
    )
    
    # Average Days Remaining
    days_remaining_list = [(item["Due Date"] - today).days for item in ROUTINE_DATA]
    avg_days_remaining = sum(days_remaining_list) / len(days_remaining_list) if days_remaining_list else 0
    
    # Backlog > 30 days (items overdue by more than 30 days)
    backlog_30_count = sum(
        1 for item in ROUTINE_DATA 
        if (item["Due Date"] - today).days < -30 and item["Status"] != "Completed"
    )
    
    # Title section
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "📊 SMU ELDIS RADAR DASHBOARD"
    title_cell.font = Font(bold=True, size=18, color=COLORS["header_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells("A2:D2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = f"Operations Control Panel | Generated: {today.strftime('%Y-%m-%d')}"
    subtitle_cell.font = Font(italic=True, size=11, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # KPI Header
    ws.merge_cells("A4:D4")
    kpi_header = ws["A4"]
    kpi_header.value = "📈 KEY PERFORMANCE INDICATORS"
    kpi_header.font = Font(bold=True, size=14, color=COLORS["header_bg"])
    
    # KPI Data
    kpis = [
        ("📋 Total Actions", total_actions, "Total number of tracked action items"),
        ("🟡 Open / In Progress", f"{open_count} / {in_progress_count}", "Items awaiting action"),
        ("✅ Completed", completed_count, "Successfully closed items"),
        ("🔴 Overdue", overdue_count, "Items past due date"),
        ("⏱️ Avg Days Remaining", f"{avg_days_remaining:.1f}", "Average days until due"),
        ("⚠️ Backlog >30 Days", backlog_30_count, "Critically overdue items"),
    ]
    
    # Headers for KPI table
    ws.cell(row=6, column=1, value="Metric")
    ws.cell(row=6, column=2, value="Value")
    ws.cell(row=6, column=3, value="Description")
    apply_header_style(ws, 6, 1, 3)
    
    # Write KPI data
    for row_idx, (metric, value, description) in enumerate(kpis, start=7):
        ws.cell(row=row_idx, column=1, value=metric)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=3, value=description)
        
        # Center the value column
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        
        # Apply borders
        for col in range(1, 4):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER
    
    # Projects Section
    ws.merge_cells("A14:D14")
    proj_header = ws["A14"]
    proj_header.value = "🏗️ PROJECT STATUS SUMMARY"
    proj_header.font = Font(bold=True, size=14, color=COLORS["header_bg"])
    
    # Project metrics
    proj_total = len(PROJECT_DATA)
    proj_open = sum(1 for p in PROJECT_DATA if p["Status"] == "Open")
    proj_in_progress = sum(1 for p in PROJECT_DATA if p["Status"] == "In Progress")
    proj_completed = sum(1 for p in PROJECT_DATA if p["Status"] == "Completed")
    
    ws.cell(row=16, column=1, value="Total Projects")
    ws.cell(row=16, column=2, value=proj_total)
    ws.cell(row=17, column=1, value="Open")
    ws.cell(row=17, column=2, value=proj_open)
    ws.cell(row=18, column=1, value="In Progress")
    ws.cell(row=18, column=2, value=proj_in_progress)
    ws.cell(row=19, column=1, value="Completed")
    ws.cell(row=19, column=2, value=proj_completed)
    
    for row in range(16, 20):
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    
    # Styling tips section
    ws.merge_cells("A21:D21")
    tips_header = ws["A21"]
    tips_header.value = "💡 CONDITIONAL FORMATTING LEGEND"
    tips_header.font = Font(bold=True, size=12, color=COLORS["header_bg"])
    
    ws.cell(row=22, column=1, value="🔴 Red")
    ws.cell(row=22, column=2, value="Overdue items")
    ws.cell(row=23, column=1, value="🟡 Yellow")
    ws.cell(row=23, column=2, value="Due within 7 days")
    ws.cell(row=24, column=1, value="🟢 Green")
    ws.cell(row=24, column=2, value="Completed items")
    
    # Set column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 15
    
    # Set row heights for better spacing
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[14].height = 25
    
    return ws


def generate_dashboard():
    """
    Main function to generate the complete SMU ELDIS Radar Dashboard.
    Creates all sheets and saves the workbook.
    """
    print("🚀 Starting SMU ELDIS Radar Dashboard Generator...")
    print("=" * 60)
    
    # Create new workbook (remove default sheet)
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create sheets in order
    print("📊 Creating Control_Panel (Dashboard)...")
    create_control_panel(wb)
    
    print("📋 Creating Routine_Table...")
    create_routine_table(wb)
    
    print("🏗️ Creating Project_Table...")
    create_project_table(wb)
    
    print("📑 Creating Lists (Data Validation)...")
    create_lists_sheet(wb)
    
    # Save workbook
    print(f"\n💾 Saving workbook as '{OUTPUT_FILE}'...")
    wb.save(OUTPUT_FILE)
    
    print("=" * 60)
    print(f"✅ Dashboard generated successfully: {OUTPUT_FILE}")
    print(f"📅 Generated on: {get_today().strftime('%Y-%m-%d')}")
    print("\n📚 Sheets created:")
    print("   1. Control_Panel - KPI Dashboard")
    print("   2. Routine_Table - SMU Action Items (Sl.No 43-51)")
    print("   3. Project_Table - GoI Tender Lifecycle")
    print("   4. Lists - Data Validation References")
    print("\n🎨 Features:")
    print("   • Conditional formatting (Red/Yellow/Green)")
    print("   • Data validation dropdowns")
    print("   • Frozen headers")
    print("   • Auto-fitted columns")
    print("   • Professional styling")
    
    return wb


if __name__ == "__main__":
    generate_dashboard()
